import os
import re
import pandas as pd
from langchain_community.document_loaders import PyPDFLoader, TextLoader
from langchain_community.embeddings import HuggingFaceEmbeddings
from langchain_community.vectorstores import Chroma
from langchain_core.documents import Document
from langchain_text_splitters import RecursiveCharacterTextSplitter

from env_config import env_bool, env_int, env_path, optional_env, require_env


# 代理配置（可选）
if optional_env("HTTP_PROXY"):
    os.environ["HTTP_PROXY"] = optional_env("HTTP_PROXY")
if optional_env("HTTPS_PROXY"):
    os.environ["HTTPS_PROXY"] = optional_env("HTTPS_PROXY")

# 配置
DOCUMENTS_DIR = str(env_path("BUILD_DOCUMENTS_DIR"))
VECTOR_DB_DIR = str(env_path("BUILD_VECTOR_DB_DIR"))
CHUNK_SIZE = env_int("DOCUMENT_CHUNK_SIZE")
CHUNK_OVERLAP = env_int("DOCUMENT_CHUNK_OVERLAP")

# Embedding 模型（CPU 上运行）
embedding_model = HuggingFaceEmbeddings(
    model_name=require_env("EMBEDDING_MODEL_NAME"),
    model_kwargs={
        'device': require_env("EMBEDDING_DEVICE"),
        'local_files_only': env_bool("EMBEDDING_LOCAL_FILES_ONLY"),
    },
    encode_kwargs={'normalize_embeddings': True}
)

# ---------- Excel 解析函数 ----------
# 这一组 Excel 解析函数的目标不是还原一个 DataFrame，而是生成 RAG 更容易理解的
# “自包含事实片段”。因此每个 chunk 都尽量同时包含：表名、sheet、父级分类、
# 指标名、真实列名和值，避免 LLM 只看到“列4: xxx”这类缺少语义的内容。
#
# 设计原则：
# 1. 不提前猜整张表属于“交叉表/横排表/普通表”，而是逐行扫描。
# 2. 每扫到一行，就根据行形态更新上下文状态：title / parent / headers。
# 3. 数据行生成 chunk 时，使用离它最近的 parent 和 headers。
# 这样一个 sheet 中有多个子表时，也能自然切换到新的子表上下文。
MONTH_OR_PERIOD_RE = re.compile(
    r"^("
    r"\d{1,2}月(?:份)?|"
    r"0?\d{1,2}月|"
    r"\d{4}年\d{1,2}月(?:份)?|"
    r"[一二三四1-4]季度|"
    r"\d{4}Q[1-4]|"
    r"Q[1-4]|"
    r"本年累计|累计数|本期数|当期数|"
    r"截至当期|截至\d{1,2}月末|截至期末|"
    r"本年累计/截至当期|月末|期末|当期|数值"
    r")$"
)

def clean_cell(value):
    """把 Excel 单元格值统一成干净字符串，便于后续规则判断。"""
    if pd.isna(value):
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return re.sub(r"\s+", " ", str(value)).strip()


def is_empty(value):
    return clean_cell(value) == ""


def normalize_label(value):
    """用于“标签类判断”的规范化：去掉所有空白，降低格式差异影响。"""
    return re.sub(r"\s+", "", clean_cell(value))


def is_note_row(label):
    # 注释/备注行通常不是业务数据，也不能当成父级标题。
    return clean_cell(label).startswith(("注：", "注:", "备注", "说明"))


def is_unit_row(row_values):
    # 例如“单位：亿元、%”。这类行只提供计量单位，不参与 chunk 生成。
    values = [clean_cell(v) for v in row_values if not is_empty(v)]
    return bool(values) and all("单位" in v for v in values)


def is_time_row(row_values):
    # 例如“时间 | 2026年”。它常常出现在真正表头上一行。
    # 当前实现会把它保存到 previous_row，供 make_headers 参考。
    return bool(row_values) and normalize_label(row_values[0]) == "时间"


def looks_like_period_header(value):
    # 判断右侧列名是否像时间/期间列，例如 1月、一季度、本年累计、Q1。
    # 它只是辅助条件；如果第一列明确是“项目/指标/名称”，不依赖这个正则。
    text = normalize_label(value)
    return bool(MONTH_OR_PERIOD_RE.match(text)) or bool(re.match(r"^\d{4}年$", text))


def looks_like_header_row(row_values):
    """判断当前行是否是表头行。

    两类行会被当成表头：
    1. 第一列明确写着“项目/指标/名称”。
    2. 右侧大多数非空单元格看起来像月份、季度、累计值等期间列。

    这比“非空最多的行就是表头”稳定，因为层级表的数据行往往非空更多。
    """
    if not row_values:
        return False
    first = normalize_label(row_values[0])
    rest = [v for v in row_values[1:] if not is_empty(v)]
    if first in {"项目", "指标", "名称"}:
        return True
    if rest and sum(1 for v in rest if looks_like_period_header(v)) >= max(1, len(rest) // 2):
        return True
    return False


def make_headers(row_values, previous_row=None):
    """生成当前子表的列名。

    row_values 是当前表头行，例如：项目 | 1月 | 2月 | 3月 | 4月。
    previous_row 是上一行，例如：时间 | 2026年 | 空 | 空 | 空。

    当前行有值时优先用当前行；当前行为空但上一行有有效值时，使用上一行补充。
    这样可以兼容多级表头，同时避免把空列退化成“列1/列2/列3”。
    """
    headers = []
    previous_row = previous_row or []
    max_len = max(len(row_values), len(previous_row))
    for idx in range(max_len):
        current = clean_cell(row_values[idx]) if idx < len(row_values) else ""
        previous = clean_cell(previous_row[idx]) if idx < len(previous_row) else ""
        if idx == 0:
            headers.append(current or previous or "项目")
        elif current:
            headers.append(current)
        elif previous and normalize_label(previous) not in {"时间", "项目", "指标", "名称"}:
            headers.append(previous)
        else:
            headers.append("")
    return headers


def is_parent_row(row_values):
    """判断是否是父级标题行。

    父级标题行的典型形态是：只有第一列有值，其他列为空。
    例如“2. 大型商业银行”。它本身没有数值，不生成 chunk；
    但会作为 parent 带入后续数据行。
    """
    values = [clean_cell(v) for v in row_values]
    non_empty_positions = [idx for idx, value in enumerate(values) if value]
    if non_empty_positions != [0]:
        return False
    label = values[0]
    if is_note_row(label) or "单位" in label:
        return False
    if normalize_label(label) in {"项目", "指标", "时间", "数值"}:
        return False
    return True


def row_has_data(row_values):
    # 数据行要求：第一列有指标名，右侧至少有一个真实值。
    if not row_values or is_empty(row_values[0]) or is_note_row(row_values[0]):
        return False
    return any(not is_empty(v) for v in row_values[1:])


def is_vertical_child_label(label):
    # 两列纵向表里，子项常写成“1、财产险”“2、人身险”“其中：xxx”。
    # 这些子项需要继承上一条主指标，例如“原保险保费收入 - 1、财产险”。
    text = normalize_label(label)
    return bool(re.match(r"^\d+[、.．]", text)) or text.startswith("其中：") or text.startswith("其中:")


def is_two_column_table(headers, row_values):
    # 判断是否像保险表那种“项目 | 数值”的纵向两列表。
    # 只有这种表才启用 vertical_parent，避免误伤月度/季度横向表。
    header_count = sum(1 for header in headers if clean_cell(header))
    value_count = sum(1 for value in row_values if clean_cell(value))
    return max(header_count, value_count) <= 2


def format_cell_value(metric, header, value):
    """格式化单元格值。

    金融监管表里比例/率常用 0.0648 表示 6.48%。
    为了让检索和问答同时命中“小数”和“百分比”两种表达，这里保留原值，
    并追加一个百分比展示，例如：0.0648（6.48%）。
    """
    value_text = clean_cell(value)
    metric_text = clean_cell(metric)
    header_text = clean_cell(header)
    metric_context = metric_text + header_text
    if any(keyword in metric_context for keyword in ("率", "占比", "比例", "净息差")):
        try:
            numeric_value = float(value_text.replace(",", ""))
        except ValueError:
            return value_text
        if -1 <= numeric_value <= 1:
            percent_text = f"{numeric_value * 100:.2f}".rstrip("0").rstrip(".")
            return f"{value_text}（{percent_text}%）"
    return value_text


def format_row_chunk(title, parent, row_label, headers, row_values, sheet_name):
    """把一行数据转换成自包含 chunk。

    示例：
    表:... | sheet:... | 2. 大型商业银行 - 总负债：1月:... | 4月:...

    这里刻意把 parent、row_label、header 和 value 放在同一个 chunk 中，
    解决父级标题和数据行被切开后 LLM 无法判断归属的问题。
    """
    data_parts = []
    for idx in range(1, min(len(headers), len(row_values))):
        header = clean_cell(headers[idx])
        value = format_cell_value(row_label, header, row_values[idx])
        if not value or not header:
            continue
        if normalize_label(header) in {"项目", "指标", "时间"}:
            continue
        data_parts.append(f"{header}:{value}")

    if not data_parts:
        return ""

    subject = clean_cell(row_label)
    if parent:
        subject = f"{clean_cell(parent)} - {subject}"

    context = []
    if title:
        context.append(f"表:{clean_cell(title)}")
    if sheet_name:
        context.append(f"sheet:{clean_cell(sheet_name)}")
    context.append(f"{subject}：" + " | ".join(data_parts))
    return " | ".join(context)


def parse_general_excel(file_path):
    """按行解析 Excel，保留真实表头，并把父级标题带入数据行。

    核心做法是把一个 sheet 当成“从上到下声明上下文”的文档：
    - 遇到标题行，记录 title；
    - 遇到父级行，记录 parent；
    - 遇到表头行，记录 headers；
    - 遇到数据行，用当前 title + parent + headers 生成 chunk。

    如果同一个 sheet 中有多个子表，后面的子表标题/表头会覆盖当前状态，
    所以后续数据行会自动进入新的子表上下文。
    """
    try:
        # header=None 很关键：不要让 pandas 自动猜表头。
        # 自动猜表头容易把真正的“项目 | 1月 | 2月”跳过，导致列名退化。
        sheets = pd.read_excel(file_path, sheet_name=None, header=None, dtype=object)
    except Exception as exc:
        print(f"Pandas 加载失败: {file_path}, 错误: {exc}")
        return []

    chunks = []
    source = os.path.basename(file_path)
    for sheet_name, df in sheets.items():
        if df is None or df.empty:
            continue
        df = df.dropna(how="all").dropna(axis=1, how="all")
        if df.empty:
            continue

        # title: 当前 sheet 或大表标题，例如“2026年银行业金融机构总资产、总负债情况表(月度)”。
        # parent: 当前子表/父级分类，例如“2. 大型商业银行”。
        # vertical_parent: 两列纵向表的隐含父级，例如“原保险保费收入”。
        # headers: 当前子表表头，例如 ["项目", "1月", "2月", "3月", "4月"]。
        # previous_row: 上一行，用于处理“时间 | 2026年”+“项目 | 1月...”这种多级表头。
        title = ""
        parent = ""
        vertical_parent = ""
        headers = []
        previous_row = []

        for row_idx, row in df.iterrows():
            row_values = [clean_cell(v) for v in row.tolist()]
            if not any(row_values):
                previous_row = row_values
                continue

            first_value = row_values[0]
            non_empty_count = sum(1 for value in row_values if value)

            # 只有一个非空单元格且还没有 title 时，通常是整张表标题。
            # 注意：后续同样只有第一列有值的行，不再当 title，而交给 parent 逻辑处理。
            if not title and non_empty_count == 1 and not is_note_row(first_value):
                title = first_value
                previous_row = row_values
                continue

            # 单位行、注释行、时间行本身不生成 chunk。
            # 时间行保存在 previous_row 中，下一行如果是“项目 | 1月...”会用它补充表头上下文。
            if is_note_row(first_value) or is_unit_row(row_values) or is_time_row(row_values):
                previous_row = row_values
                continue

            # 识别到新表头时，更新 headers。
            # 这也是支持同一 sheet 多个子表的关键：每个子表可以有自己的表头。
            if looks_like_header_row(row_values):
                headers = make_headers(row_values, previous_row)
                previous_row = row_values
                continue

            # 父级行不含数值，但它定义了后续数据行的归属。
            # 例如“2. 大型商业银行”后面的“总负债”都应带上这个 parent。
            if is_parent_row(row_values):
                parent = first_value
                vertical_parent = ""
                previous_row = row_values
                continue

            if row_has_data(row_values):
                # 极端情况下没有识别到表头，也给一个兜底列名，保证不会丢行。
                # 正常月度/季度表会在前面的 looks_like_header_row 分支里拿到真实表头。
                if not headers:
                    headers = make_headers(["项目"] + [f"值{i}" for i in range(1, len(row_values))], previous_row)
                row_parent = parent

                # 两列纵向表中，子项需要继承上一条主指标。
                # 例如：
                # 原保险保费收入 | 27329.23
                # 1、财产险      | 4929.38
                # 应输出为“原保险保费收入 - 1、财产险”，否则会和其他“1、财产险”混淆。
                if is_two_column_table(headers, row_values) and is_vertical_child_label(first_value):
                    row_parent = vertical_parent or parent
                chunk_text = format_row_chunk(title, row_parent, first_value, headers, row_values, sheet_name)
                if chunk_text:
                    chunks.append(
                        Document(
                            page_content=chunk_text,
                            metadata={
                                "source": source,
                                "sheet": str(sheet_name),
                                "row_index": int(row_idx),
                                "parent": clean_cell(row_parent),
                                "type": "excel_general",
                            },
                        )
                    )

                # 在两列纵向表中，非子项数据行会成为后续子项的 vertical_parent。
                # 月度/季度横向表不会进入这里，因为它不是 two-column table。
                if is_two_column_table(headers, row_values) and not is_vertical_child_label(first_value):
                    vertical_parent = first_value

            # 每一轮都记录上一行，供下一行处理多级表头或上下文。
            previous_row = row_values

    return chunks


def load_excel(file_path):
    """统一使用结构感知解析，避免错误路由导致表头/父级丢失。"""
    chunks = parse_general_excel(file_path)
    if chunks:
        return chunks
    return load_excel_fallback(file_path)

def load_excel_fallback(file_path):
    """Pandas 后备解析，读取所有 sheet 并逐行转文本"""
    try:
        sheets = pd.read_excel(file_path, sheet_name=None, dtype=object)
    except Exception as exc:
        print(f"Pandas 加载失败: {file_path}, 错误: {exc}")
        return []

    chunks = []
    source = os.path.basename(file_path)
    for sheet_name, df in sheets.items():
        if df is None or df.empty:
            continue
        df = df.dropna(how="all")
        if df.empty:
            continue

        # 用列索引生成列名
        columns = []
        for idx, col in enumerate(df.columns):
            col_name = str(col).strip()
            if not col_name or col_name.lower().startswith("unnamed:"):
                col_name = f"column_{idx + 1}"
            columns.append(col_name)

        for row_idx, row in df.iterrows():
            row_parts = []
            for col_name, val in zip(columns, row.tolist()):
                if pd.isna(val):
                    continue
                val_text = str(val).strip()
                if val_text:
                    row_parts.append(f"{col_name}: {val_text}")
            if row_parts:
                chunks.append(
                    Document(
                        page_content=" | ".join(row_parts),
                        metadata={
                            "source": source,
                            "sheet": str(sheet_name),
                            "row_index": int(row_idx),
                            "type": "pandas_fallback"
                        }
                    )
                )
    return chunks

def load_with_eparse(file_path):
    try:
        from eparse.core import get_df_from_file
    except ImportError:
        print("eparse 未安装，请运行: pip install eparse")
        return []
    chunks = []
    tables = get_df_from_file(file_path)
    for table_idx, (df, metadata) in enumerate(tables):
        if df.empty:
            continue
        sheet_name = metadata.get('sheet_name', 'Unknown')
        table_range = metadata.get('range', '')
        for _, row in df.iterrows():
            row_parts = []
            for col, val in row.items():
                if pd.notna(val) and str(val).strip():
                    row_parts.append(f"{col}: {val}")
            if row_parts:
                chunks.append(
                    Document(
                        page_content=" | ".join(row_parts),
                        metadata={
                            "source": os.path.basename(file_path),
                            "sheet": sheet_name,
                            "table_index": table_idx,
                            "range": table_range,
                            "type": "eparse"
                        }
                    )
                )
    return chunks

def load_with_ks_parser(file_path):
    try:
        from ks_xlsx_parser import parse_xlsx
    except ImportError:
        print("ks-xlsx-parser 未安装，请运行: pip install ks-xlsx-parser")
        return []
    result = parse_xlsx(file_path)
    chunks = []
    for sheet in result.get("sheets", []):
        sheet_name = sheet.get("name", "Unknown")
        for table_idx, table in enumerate(sheet.get("tables", [])):
            headers = table.get("headers", [])
            rows = table.get("rows", [])
            table_range = table.get("range", "")
            for row_idx, row in enumerate(rows):
                row_data = []
                for i, val in enumerate(row):
                    if i < len(headers) and val and str(val).strip():
                        row_data.append(f"{headers[i]}: {val}")
                if row_data:
                    cell_refs = table.get("cell_refs", [])
                    ref_str = " | ".join(cell_refs[row_idx]) if cell_refs and row_idx < len(cell_refs) else ""
                    chunks.append(
                        Document(
                            page_content=" | ".join(row_data),
                            metadata={
                                "source": os.path.basename(file_path),
                                "sheet": sheet_name,
                                "table_index": table_idx,
                                "row_index": row_idx,
                                "range": table_range,
                                "cell_refs": ref_str,
                                "type": "ks_parser"
                            }
                        )
                    )
    return chunks

def has_complex_structure(file_path):
    try:
        from openpyxl import load_workbook
        wb = load_workbook(file_path, data_only=False)
        sheet = wb.active
        if sheet.merged_cells:
            return True
        for row in sheet.iter_rows(max_row=100):
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    return True
    except:
        pass
    return False

# ---------- 其他文档加载 ----------
def load_markdown(file_path):
    loader = TextLoader(file_path, encoding='utf-8')
    docs = loader.load()
    for doc in docs:
        doc.metadata["source"] = os.path.basename(file_path)
        doc.metadata["type"] = "markdown"
    return docs

def load_pdf(file_path):
    loader = PyPDFLoader(file_path)
    docs = loader.load()
    for doc in docs:
        doc.metadata["source"] = os.path.basename(file_path)
        doc.metadata["type"] = "pdf"
    return docs

def load_document(file_path):
    ext = os.path.splitext(file_path)[1].lower()
    if ext == '.pdf':
        return load_pdf(file_path)
    elif ext in ['.xls', '.xlsx']:
        return load_excel(file_path)
    elif ext == '.md':
        return load_markdown(file_path)
    else:
        raise ValueError(f"Unsupported file type: {ext}")

# ---------- 构建向量库 ----------
def build_vector_store():
    # 检查目录是否存在
    if not os.path.exists(DOCUMENTS_DIR):
        print(f"错误：目录 {DOCUMENTS_DIR} 不存在")
        return

    # 列出支持的文件
    supported_exts = ('.pdf', '.xls', '.xlsx', '.md')
    files = [f for f in os.listdir(DOCUMENTS_DIR) if f.lower().endswith(supported_exts)]
    if not files:
        print(f"警告：在 {DOCUMENTS_DIR} 中没有找到支持的文档文件")
        print("支持的文件类型：", supported_exts)
        return

    all_docs = []
    for filename in files:
        file_path = os.path.join(DOCUMENTS_DIR, filename)
        print(f"正在加载: {file_path}")
        docs = load_document(file_path)
        if not docs:
            print(f"  加载 {filename} 返回空，跳过")
            continue
        # 对 PDF/MD 切分，Excel 已经按行切分，不再切分
        if filename.endswith(('.pdf', '.md')):
            splitter = RecursiveCharacterTextSplitter(
                chunk_size=CHUNK_SIZE,
                chunk_overlap=CHUNK_OVERLAP,
                separators=["\n\n", "\n", "。", "！", "？", "；", "，", " ", ""]
            )
            docs = splitter.split_documents(docs)
        all_docs.extend(docs)
        print(f"  -> 生成了 {len(docs)} 个片段")

    if not all_docs:
        print("错误：没有任何文档片段被生成，请检查文档内容或解析逻辑。")
        return

    print(f"总计片段数: {len(all_docs)}")
    # 构建向量库
    vectordb = Chroma.from_documents(
        documents=all_docs,
        embedding=embedding_model,
        persist_directory=VECTOR_DB_DIR
    )
    vectordb.persist()
    print(f"向量库已保存至 {VECTOR_DB_DIR}")

if __name__ == "__main__":
    if not os.path.exists(DOCUMENTS_DIR):
        os.makedirs(DOCUMENTS_DIR)
        print(f"Please put your documents in {DOCUMENTS_DIR} and run again.")
    else:
        build_vector_store()
