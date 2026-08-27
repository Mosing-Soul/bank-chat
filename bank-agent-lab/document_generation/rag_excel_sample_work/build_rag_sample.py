from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


OUTPUT_DIR = Path(__file__).resolve().parent.parent / "rag_excel_sample_v1"
OUTPUT_FILE = OUTPUT_DIR / "华辰银行零售客户经理知识库小样_V1.0.xlsx"

NAVY = "17365D"
BLUE = "2F75B5"
LIGHT_BLUE = "D9EAF7"
PALE_BLUE = "EEF5FB"
GOLD = "D6A84B"
LIGHT_GOLD = "FFF2CC"
GREEN = "70AD47"
LIGHT_GREEN = "E2F0D9"
RED = "C00000"
LIGHT_RED = "FCE4D6"
GRAY = "E7E6E6"
WHITE = "FFFFFF"
TEXT = "1F2937"
THIN = Side(style="thin", color="C9D2DC")
MEDIUM = Side(style="medium", color=NAVY)


qa_rows = [
    ["QA-RB-001", "客户分层", "金穗级", "金穗级客户需要达到什么条件？", "金穗级门槛；客户升级条件", "个人客户", "客户最近连续3个月月日均综合金融资产达到50万元，且客户身份资料完整、风险评级有效。该标准为华辰银行内部参考演示口径，不代表任何真实银行。", "连续3个月达标；KYC有效", "客户有效身份证件、联系方式", "客户经理工作台自动识别，次月5日前更新", "不得承诺永久保级；异常资产需核实来源", "客户质疑资产计算或系统未更新", "SOP-RB-001", "内部参考", "V1.0"],
    ["QA-RB-002", "客户分层", "白金级", "白金级客户的资产门槛是多少？", "白金客户标准；高净值客户门槛", "个人客户", "最近连续3个月月日均综合金融资产达到200万元可进入白金级候选名单，经系统校验后生效。本条为内部参考口径。", "连续3个月达标", "无需额外申请材料", "系统批量认定", "不得用临时转入资金诱导冲量", "名单异常或客户提出异议", "SOP-RB-001", "内部参考", "V1.0"],
    ["QA-RB-003", "客户分层", "降级保护", "客户资产短期下降会马上降级吗？", "保级期多久；资产下降怎么办", "个人客户", "不会立即降级。华辰银行内部参考规则设置60天观察期；观察期结束仍未恢复标准的，次月调整等级。涉嫌异常交易或账户风险管控的除外。", "非风险事件导致的自然波动", "无", "系统观察名单", "风险管控不适用保级承诺", "客户要求人工保级", "SOP-RB-001", "内部参考", "V1.0"],
    ["QA-AC-001", "账户服务", "借记卡开户", "客户新开借记卡需要带什么？", "办银行卡材料；开户证件", "年满18周岁的境内个人", "客户本人携带有效居民身份证原件，提供本人实名手机号，并配合完成开户用途、职业、常住地址和受益所有人等身份核实。特殊情形按柜面要求补充辅助证明。", "本人办理；身份可核验", "身份证原件、实名手机号；必要时辅助证明", "网点柜面", "不得以营销需要简化客户身份识别", "身份核验失败、疑似冒名或异常开户", "SOP-AC-001", "行业通行内部参考", "V1.0"],
    ["QA-AC-002", "账户服务", "账户分类", "一类户和二类户有什么区别？", "I类II类账户区别；二类户限制", "个人客户", "一类户通常作为全功能结算账户；二类户功能和交易限额受监管及本行渠道规则限制。具体限额应以办理当日系统提示为准，客户经理不得口头承诺固定限额。", "账户状态正常", "无", "网点或支持的电子渠道", "限额会随监管和风控策略调整", "客户要求突破限额", "SOP-AC-001", "公开规则概括+内部参考口径", "V1.0"],
    ["QA-AC-003", "账户服务", "手机号变更", "客户换手机号后怎么修改预留手机号？", "改手机；预留号码变更", "个人客户", "原则上由客户本人持有效身份证件和银行卡到网点办理；若电子渠道支持且客户通过增强认证，可按页面流程办理。涉及高风险账户时仅限柜面核验。", "账户可正常验证", "身份证件、银行卡、新实名手机号", "网点；符合条件时电子渠道", "不得代客户接收验证码", "原手机号停用且认证失败、账户被管控", "SOP-AC-002", "行业通行内部参考", "V1.0"],
    ["QA-DP-001", "存款业务", "大额存单", "大额存单可以提前支取吗？", "提前取大额存单；提前支取利息", "持有本行大额存单的个人客户", "是否允许提前支取、可支取次数及计息方式，以具体产品说明书和系统展示为准。办理前必须向客户说明提前支取可能造成利息损失。", "产品支持提前支取", "身份证件、账户介质", "网点或产品支持的电子渠道", "不得将历史产品规则套用于当前产品", "系统与产品说明书不一致", "SOP-DP-001", "产品规则内部参考", "V1.0"],
    ["QA-DP-002", "存款业务", "定期存款", "定期存款到期会自动续存吗？", "自动转存；定期到期怎么办", "个人客户", "取决于开户时选择。约定自动转存的，通常按到期日对应期限的挂牌利率续存；未约定的，按产品协议处理。应查询原交易回单或系统约定，不凭经验回答。", "账户及存单状态正常", "无", "客户经理工作台查询后答复", "利率以到期日实际规则为准", "历史纸质存单无法查询", "SOP-DP-001", "行业通行内部参考", "V1.0"],
    ["QA-LN-001", "个人贷款", "消费贷款", "申请个人消费贷通常需要哪些材料？", "消费贷材料；贷款申请资料", "有稳定收入的个人客户", "基础材料包括身份证明、婚姻及居住信息、收入或还款能力证明、贷款用途材料和征信授权；系统预审批不代表最终获批，具体以客户情况和产品要求为准。", "年龄、征信、收入等满足产品准入", "身份证明、收入证明、用途证明、征信授权", "客户经理受理后提交线上审批", "严禁包装用途或承诺审批结果", "用途不清、收入异常、征信异议", "SOP-LN-001", "行业通行内部参考", "V1.0"],
    ["QA-LN-002", "个人贷款", "经营贷款", "个体工商户申请经营贷要准备什么？", "经营贷材料；个体户贷款", "持续经营的个体工商户", "除个人身份材料外，通常需营业执照、经营流水、经营场所或上下游证明、贷款用途材料及征信授权。不同担保方式还需补充抵押物或保证人资料。", "真实经营且经营期限满足产品规则", "营业执照、经营流水、用途证明、征信授权", "网点或线上预约后尽调", "不得将消费用途包装成经营用途", "经营真实性存疑或现金流水占比异常", "SOP-LN-002", "行业通行内部参考", "V1.0"],
    ["QA-LN-003", "个人贷款", "住房按揭", "按揭贷款审批一般要多久？", "房贷多久批；按揭审批时效", "符合本行按揭准入的购房客户", "材料齐全后，华辰银行内部服务参考时限为5个工作日内完成内部审批；抵押登记和放款时间还受交易、登记机关及额度安排影响，不得向客户承诺固定放款日。", "材料齐全且外部查询正常", "购房、收入、首付款及征信相关材料", "网点受理", "审批时效不等于放款时效", "交易临近交割或客户要求承诺日期", "SOP-LN-003", "内部参考", "V1.0"],
    ["QA-LN-004", "个人贷款", "提前还款", "贷款提前还款有没有违约金？", "提前还贷费用；提前结清", "存量个人贷款客户", "是否收取违约金以及预约期限，应以客户贷款合同、产品规则和系统试算结果为准。客户经理应先查询合同，不得统一回答“有”或“没有”。", "贷款状态正常", "身份证件、还款账户；按合同补充", "合同约定渠道", "以合同约定为优先依据", "合同条款不清或系统无法试算", "SOP-LN-004", "合同口径", "V1.0"],
    ["QA-WM-001", "财富管理", "风险测评", "客户风险测评过期还能买理财吗？", "测评失效；风险评级过期", "拟购买投资产品的个人客户", "风险测评失效后不得直接完成产品购买，应先由客户本人重新测评并确认结果。客户经理不得代答、诱导选择或修改客户答案。", "客户具备民事行为能力并本人操作", "有效身份证明", "网点或合规电子渠道", "测评结果应真实反映客户情况", "客户拒绝测评或疑似被代操作", "SOP-WM-001", "适当性通行要求", "V1.0"],
    ["QA-WM-002", "财富管理", "适当性匹配", "稳健型客户能买高风险产品吗？", "风险不匹配能否购买；越级购买", "已完成有效风险测评的客户", "原则上应购买与其风险承受能力相匹配的产品。发生不匹配时，必须按适用法规、产品规则和系统控制执行额外提示、确认或禁止销售，不得规避系统限制。", "风险测评有效", "无", "系统适当性校验", "不得通过重测诱导客户提高风险等级", "客户坚持购买且系统拦截", "SOP-WM-001", "适当性通行要求", "V1.0"],
    ["QA-WM-003", "财富管理", "理财赎回", "理财赎回后资金什么时候到账？", "赎回到账；理财取出多久到账", "持有本行代销或发行理财产品的客户", "到账时间取决于产品开放规则、赎回确认日和清算安排，应以产品说明书及交易页面预计到账日为准。节假日可能顺延。", "产品处于可赎回状态", "无", "原购买渠道", "不得用其他产品到账经验代替本产品规则", "超过预计到账日仍未到账", "SOP-WM-002", "产品规则内部参考", "V1.0"],
    ["QA-WM-004", "财富管理", "基金定投", "基金定投可以随时暂停吗？", "暂停定投；取消基金扣款", "已签约基金定投的客户", "通常可在约定扣款日前通过原签约渠道暂停或终止，具体截止时间以系统提示和基金销售规则为准。已进入扣款处理的当期计划可能无法撤销。", "签约状态正常", "无", "手机银行或网点", "暂停定投不等于赎回已持有份额", "当期已扣款但客户要求撤销", "SOP-WM-002", "产品规则内部参考", "V1.0"],
    ["QA-IN-001", "保险代销", "双录", "哪些保险销售场景需要双录？", "保险双录；录音录像要求", "购买本行代销保险的客户", "应根据客户年龄、销售渠道、产品类型及现行监管和本行制度判断。系统提示需双录的，必须在规定场所和流程内完成，不得拆分交易或变更渠道规避。", "客户本人投保", "身份证明及投保资料", "网点双录专区或合规远程渠道", "双录未完成不得进入后续销售确认", "客户拒绝录制或设备故障", "SOP-IN-001", "监管概括+内部参考流程", "V1.0"],
    ["QA-IN-002", "保险代销", "犹豫期", "保险犹豫期从哪天开始算？", "退保犹豫期；保险冷静期", "已投保客户", "犹豫期起算和期限以保险合同、回执签收及保险公司规则为准。客户经理应查阅具体保单，不得只按产品名称推断。", "保单已生效", "保险合同或电子保单", "联系保险公司或本行售后协助", "犹豫期退保也可能存在合同约定费用", "临近截止日或签收日期有争议", "SOP-IN-002", "合同口径", "V1.0"],
    ["QA-KY-001", "客户尽调", "KYC更新", "客户身份证过期了还能办理业务吗？", "证件过期；身份证失效", "证件已过期的存量客户", "可办理范围取决于业务类型、过期时长和账户风险状态。应先提示客户更新证件；涉及新增产品、资金划转或高风险业务时，系统可能限制办理。", "本人身份可核实", "新有效身份证件", "网点或支持证件更新的电子渠道", "不得绕过证件有效期校验", "客户无法及时更新且有紧急资金需求", "SOP-KY-001", "KYC通行要求", "V1.0"],
    ["QA-KY-002", "客户尽调", "职业信息", "为什么还要问客户职业和收入来源？", "KYC为什么问职业；资金来源", "办理开户、贷款或高风险业务的客户", "银行需要履行客户身份识别、风险管理和反洗钱义务。相关信息用于判断交易与客户身份、职业及资金来源是否匹配，并非用于无关营销。", "依法依规收集", "客户说明；必要时提供证明", "受理环节", "只收集履职所必需的信息", "客户拒绝提供关键身份信息", "SOP-KY-001", "反洗钱要求概括", "V1.0"],
    ["QA-AML-001", "反洗钱", "异常交易", "客户突然转入大额资金，客户经理应该怎么做？", "大额入账处理；异常资金核实", "交易行为与历史明显不一致的客户", "不得仅凭金额直接认定异常，也不得协助规避监测。应了解资金来源、交易背景和用途，留存合理说明及必要证明，并按内部预警流程提交核查。", "交易触发系统预警或人工关注", "资金来源、交易背景相关证明", "客户经理工作台预警任务", "不得提示客户拆分交易", "客户拒绝说明或材料明显矛盾", "SOP-AML-001", "反洗钱通行要求", "V1.0"],
    ["QA-AML-002", "反洗钱", "可疑报告", "能告诉客户银行已经报送可疑交易吗？", "可疑交易能否告知；反洗钱保密", "涉及可疑交易监测的客户", "不得向客户或无关人员泄露可疑交易报告、内部分析结论和报送情况。对客户的业务解释应使用经批准的服务口径。", "涉及内部监测或报告", "无", "内部合规流程", "严格遵守保密和知情范围控制", "客户投诉或要求书面说明", "SOP-AML-001", "反洗钱要求概括", "V1.0"],
    ["QA-FR-001", "反欺诈", "账户管控", "客户账户突然不能转账怎么办？", "转账失败；账户被限制", "账户交易受限客户", "先核验客户身份并查询可对客展示的限制类型，再告知办理渠道和材料。不得透露内部风险模型、名单来源或具体监测规则。司法冻结等情形按有权机关协助规定处理。", "本人咨询", "身份证件；按限制原因补充", "网点或客服协同", "不得承诺立即解除", "涉及司法冻结、疑似诈骗或客户情绪激烈", "SOP-FR-001", "行业通行内部参考", "V1.0"],
    ["QA-FR-002", "反欺诈", "被骗处置", "客户说刚被骗并已经转账，第一步做什么？", "被骗汇款；紧急止付", "疑似遭遇电信网络诈骗的客户", "立即引导客户停止继续转账，保存聊天和交易证据，并尽快报警；同时通过本行紧急渠道申请账户保护和交易核查。能否止付或追回取决于资金状态，不得承诺结果。", "客户能提供基本交易信息", "身份信息、交易时间、金额、收款账户", "客服、网点及公安机关", "争分夺秒但不得保证追回", "仍在受骗控制、持续转账或人身安全风险", "SOP-FR-002", "行业通行处置", "V1.0"],
    ["QA-CS-001", "投诉服务", "投诉受理", "客户要正式投诉应该怎么登记？", "投诉登记；客户不满意", "提出服务异议的客户", "记录客户身份、联系方式、争议业务、发生时间、核心诉求和期望解决方式，生成投诉编号并告知查询渠道。不得以“建议”名义规避投诉登记。", "客户明确表达不满或提出处理诉求", "客户陈述及相关业务凭证", "网点、客服或线上渠道", "不得要求客户撤诉作为解决条件", "重大舆情、人身威胁或群体事件苗头", "SOP-CS-001", "消保通行要求", "V1.0"],
    ["QA-CS-002", "投诉服务", "处理时效", "投诉多久能有回复？", "投诉回复时间；几天处理", "已登记投诉的客户", "受理时应告知预计处理时间。华辰银行内部服务参考时限为简单投诉3个工作日内反馈，复杂投诉15日内反馈阶段性进展；法定或监管时限优先适用。", "投诉已成功登记", "投诉编号", "原受理渠道", "复杂案件不得无故长期挂起", "超过承诺时间或客户要求监管投诉", "SOP-CS-001", "内部参考", "V1.0"],
    ["QA-MK-001", "营销合规", "收益表述", "可以告诉客户某款理财肯定赚钱吗？", "理财保本吗；承诺收益", "投资产品潜在客户", "不可以。不得承诺保本保收益，不得使用“稳赚”“零风险”等误导性表述。应说明产品风险、业绩比较基准不代表实际收益，并引导客户阅读产品文件。", "完成必要的产品介绍前置步骤", "产品说明书和风险揭示书", "合规销售渠道", "营销材料须使用审批版本", "客户要求书面收益承诺", "SOP-MK-001", "销售合规通行要求", "V1.0"],
    ["QA-MK-002", "营销合规", "外呼", "客户经理可以用个人微信发产品海报吗？", "微信营销；个人号发产品", "存量客户营销", "仅可使用本行批准的渠道和物料。未经审批的海报、个人加工的收益对比及含客户隐私的截图不得发送；是否允许使用工作微信以本行制度为准。", "客户已同意接收相应营销信息", "审批版营销材料", "本行批准渠道", "尊重退订和拒收选择", "物料来源不明或客户投诉骚扰", "SOP-MK-001", "营销合规内部参考", "V1.0"],
    ["QA-PR-001", "隐私保护", "信息查询", "家属能不能查询客户账户余额？", "配偶查余额；家人查询账户", "客户家属或关系人", "未经客户本人有效授权或法律明确规定，不得向家属披露账户余额、交易明细、产品持仓等客户信息。婚姻或亲属关系本身不等于查询授权。", "核验查询人身份和授权依据", "本人授权文件或有权机关法律文书", "网点合规受理", "最小必要披露", "授权真伪存疑或家庭纠纷", "SOP-PR-001", "隐私保护通行要求", "V1.0"],
    ["QA-PR-002", "隐私保护", "资料传输", "客户可以把身份证照片发给客户经理吗？", "微信发身份证；传客户资料", "需要提交业务材料的客户", "应优先引导客户通过本行批准的安全渠道上传。不得要求客户发送到个人邮箱或个人社交账号；已误收的，应按本行信息安全流程处置，不得自行留存或转发。", "确有业务材料收集需要", "按业务最小必要范围", "官方上传渠道或网点", "禁止在个人设备长期保存", "材料已泄露或发送给错误人员", "SOP-PR-001", "隐私保护内部参考", "V1.0"],
    ["QA-OP-001", "服务运营", "上门服务", "行动不便的老人能申请上门服务吗？", "老人上门；特殊客户服务", "因高龄、疾病或残障确有困难的客户", "可登记特殊服务申请，由网点评估业务是否支持上门见证或其他便利安排。上门服务不等于放宽身份核验、真实意愿和授权要求。", "客户本人有真实业务需求", "身份证明、情况说明；按业务补充", "联系开户网点", "至少双人服务并完整留痕（内部作业要求）", "疑似胁迫、认知能力存疑或代理争议", "SOP-OP-001", "行业通行内部参考", "V1.0"],
]


def set_title(ws, text, end_col, subtitle):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_col)
    c = ws.cell(1, 1, text)
    c.font = Font(name="微软雅黑", size=16, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=end_col)
    c = ws.cell(2, 1, subtitle)
    c.font = Font(name="微软雅黑", size=9, italic=True, color="5B6573")
    c.fill = PatternFill("solid", fgColor="F2F5F8")
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 32


def style_header(ws, row, start_col, end_col, fill=BLUE):
    for c in ws.iter_cols(min_col=start_col, max_col=end_col, min_row=row, max_row=row):
        cell = c[0]
        cell.font = Font(name="微软雅黑", size=10, bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)


def style_body(ws, min_row, max_row, max_col):
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.font = Font(name="微软雅黑", size=9, color=TEXT)
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            cell.border = Border(bottom=THIN)
        if row[0].row % 2 == 0:
            for cell in row:
                cell.fill = PatternFill("solid", fgColor="F8FAFC")


def merge_same_values(ws, col, start_row, end_row):
    group_start = start_row
    current = ws.cell(start_row, col).value
    for row in range(start_row + 1, end_row + 2):
        value = ws.cell(row, col).value if row <= end_row else object()
        if value != current:
            if row - group_start > 1:
                ws.merge_cells(start_row=group_start, start_column=col, end_row=row - 1, end_column=col)
                ws.cell(group_start, col).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                ws.cell(group_start, col).fill = PatternFill("solid", fgColor=PALE_BLUE)
                ws.cell(group_start, col).font = Font(name="微软雅黑", size=10, bold=True, color=NAVY)
            group_start = row
            current = value


def add_qa_sheet(wb):
    ws = wb.active
    ws.title = "客户经理QA"
    set_title(ws, "华辰银行零售客户经理常见问答库（内部培训版）", 15,
              "用途：RAG导入与检索测试；全部机构、产品及内部时效均为演示口径。表内保留多层表头和纵向合并分类。")
    groups = [(1, 3, "问题索引"), (4, 5, "客户常见表达"), (6, 11, "标准服务口径"), (12, 15, "治理与追溯")]
    for start, end, label in groups:
        ws.merge_cells(start_row=3, start_column=start, end_row=3, end_column=end)
        ws.cell(3, start, label)
    style_header(ws, 3, 1, 15, NAVY)
    headers = ["项目", "业务大类", "业务子类", "标准问题", "同义问法", "适用客户", "标准答案", "前置条件", "所需材料", "办理渠道/时效", "风险提示", "升级/转人工条件", "关联SOP编号", "口径属性", "版本号"]
    ws.append(headers)
    style_header(ws, 4, 1, 15, BLUE)
    ws.row_dimensions[4].height = 42
    for row in qa_rows:
        ws.append(row)
    end = 4 + len(qa_rows)
    style_body(ws, 5, end, 15)
    merge_same_values(ws, 2, 5, end)
    widths = [14, 13, 14, 28, 24, 21, 58, 25, 28, 28, 30, 28, 16, 20, 10]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    for r in range(5, end + 1):
        ws.row_dimensions[r].height = 72
    ws.freeze_panes = "D5"
    ws.auto_filter.ref = f"A4:O{end}"
    ws.sheet_view.showGridLines = False
    ws.print_title_rows = "1:4"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    dv = DataValidation(type="list", formula1='"内部参考,行业通行内部参考,公开规则概括+内部参考口径,适当性通行要求,合同口径,KYC通行要求,反洗钱通行要求,销售合规通行要求,隐私保护通行要求"')
    ws.add_data_validation(dv)
    dv.add(f"N5:N{end}")


def add_matrix_sheet(wb):
    ws = wb.create_sheet("准入与材料矩阵")
    set_title(ws, "零售业务准入与材料矩阵（内部参考）", 12,
              "说明：同一Sheet内包含多个子表；最低层表头使用完整语义列名，上层保留真实业务常见的合并分组。")
    # Subtable 1
    ws.merge_cells("A3:L3"); ws["A3"] = "子表一：个人贷款准入及材料要求"
    style_header(ws, 3, 1, 12, NAVY)
    for rng, value in [("A4:C4", "业务识别"), ("D4:F4", "基础准入"), ("G4:J4", "材料要求"), ("K4:L4", "处理规则")]:
        ws.merge_cells(rng); ws[rng.split(":")[0]] = value
    style_header(ws, 4, 1, 12, BLUE)
    headers = ["项目", "产品类型", "适用客群", "年龄范围(内部参考)", "经营/工作要求", "信用要求", "身份证明", "收入/流水材料", "用途材料", "担保材料", "标准处理时效(内部参考)", "关键控制点"]
    ws.append(headers); style_header(ws, 5, 1, 12, "5B9BD5")
    data = [
        ["消费贷款", "个人消费信用贷", "受薪人士", "22-60周岁", "当前单位连续工作满6个月", "无当前逾期，综合评分通过", "有效身份证件", "工资流水或个税/社保记录", "装修、教育等真实用途材料", "信用方式无", "材料齐全后3个工作日", "不得包装用途；预审批不等于授信"],
        ["经营贷款", "个人经营贷", "个体工商户/小微企业主", "22-65周岁", "持续经营满12个月", "个人及经营主体信用正常", "身份证件、婚姻证明", "经营流水、纳税或进销存材料", "采购合同、订单等经营用途材料", "按抵押/保证方式补充", "尽调完成后5个工作日", "核实经营真实性与资金回流"],
        ["住房贷款", "一手房按揭", "符合购房政策的自然人", "18-70周岁且贷款到期年龄合规", "具备稳定还款来源", "征信及负债符合政策", "身份证件、婚姻证明", "收入证明、银行流水", "购房合同、首付款证明", "所购住房抵押", "材料齐全后5个工作日审批", "不得承诺放款日；核实首付款来源"],
        ["住房贷款", "二手房按揭", "二手住房买受人", "18-70周岁且贷款到期年龄合规", "具备稳定还款来源", "征信及负债符合政策", "身份证件、婚姻证明", "收入证明、银行流水", "买卖合同、首付款证明", "交易住房抵押", "评估完成后5个工作日审批", "核验交易真实性和评估价值"],
    ]
    for row in data: ws.append(row)
    style_body(ws, 6, 9, 12)
    ws.merge_cells("A10:L10"); ws["A10"] = "注：以上年龄、时效和产品条件全部为内部参考演示值；实际办理以具体产品制度、监管政策及系统校验为准。"
    ws["A10"].fill = PatternFill("solid", fgColor=LIGHT_GOLD); ws["A10"].font = Font(name="微软雅黑", size=9, bold=True, color="7F6000"); ws["A10"].alignment = Alignment(wrap_text=True, vertical="center")
    # Subtable 2
    ws.merge_cells("A12:L12"); ws["A12"] = "子表二：客户身份识别材料分层清单"
    style_header(ws, 12, 1, 12, NAVY)
    for rng, value in [("A13:C13", "识别场景"), ("D13:F13", "基础材料"), ("G13:I13", "增强核实"), ("J13:L13", "处置要求")]:
        ws.merge_cells(rng); ws[rng.split(":")[0]] = value
    style_header(ws, 13, 1, 12, BLUE)
    headers2 = ["项目", "风险等级", "典型触发场景", "有效身份证件", "联系方式核验", "职业/地址信息", "资金来源证明", "交易背景证明", "其他辅助材料", "允许办理范围", "升级路径", "留痕要求"]
    ws.append(headers2); style_header(ws, 14, 1, 12, "5B9BD5")
    data2 = [
        ["常规识别", "低/中", "开户、资料更新、常规产品签约", "必需", "实名手机号", "客户如实申报", "按业务需要", "按业务需要", "系统提示时补充", "系统校验通过后办理", "身份存疑转增强核实", "影像、问询结果和系统记录"],
        ["增强识别", "中高", "交易突增、异地异常、高风险产品", "必需并联网核查", "交叉验证", "必要时提供证明", "收入、资产出售、经营等证明", "合同、发票、交易对手说明", "按预警任务补充", "核实完成前限制高风险业务", "提交网点反洗钱岗", "完整记录判断依据和材料"],
        ["拒绝/中止", "高", "疑似冒名、材料矛盾、拒绝说明", "无法有效核验", "无法核验", "拒绝或明显矛盾", "拒绝提供", "无法合理说明", "存在伪造嫌疑", "中止或拒绝业务", "报告运营主管及合规人员", "记录客观事实，不向客户泄露监测规则"],
    ]
    for row in data2: ws.append(row)
    style_body(ws, 15, 17, 12)
    for col, width in enumerate([17, 18, 24, 20, 19, 21, 24, 26, 22, 26, 24, 28], 1): ws.column_dimensions[get_column_letter(col)].width = width
    for r in list(range(6, 10)) + list(range(15, 18)): ws.row_dimensions[r].height = 66
    ws.freeze_panes = "D6"; ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape"; ws.page_setup.fitToWidth = 1; ws.sheet_properties.pageSetUpPr.fitToPage = True


def add_sop_sheet(wb):
    ws = wb.create_sheet("办理流程SOP")
    set_title(ws, "零售客户经理办理流程 SOP 矩阵（内部参考）", 11,
              "流程采用阶段→步骤的层级结构；阶段单元格纵向合并，模拟真实运营管理台账。")
    for rng, value in [("A3:C3", "流程定位"), ("D3:F3", "执行要求"), ("G3:H3", "输入输出"), ("I3:K3", "风险与服务")]:
        ws.merge_cells(rng); ws[rng.split(":")[0]] = value
    style_header(ws, 3, 1, 11, NAVY)
    headers = ["项目", "流程阶段", "责任角色", "操作动作", "系统/渠道(内部参考)", "标准时限(内部参考)", "输入材料", "输出/留痕", "关键控制点", "异常分支", "客户沟通口径"]
    ws.append(headers); style_header(ws, 4, 1, 11, BLUE)
    rows = [
        ["SOP-LN-001-01", "一、需求识别", "客户经理", "确认贷款用途、金额、期限和还款来源，完成初步适配判断", "客户经理工作台", "当场", "客户口述及基础身份信息", "需求记录", "不得诱导虚构用途", "用途不清则暂停推荐", "先了解您的真实资金用途，再匹配合适产品"],
        ["SOP-LN-001-02", "一、需求识别", "客户经理", "说明产品基本条件、费用、风险及审批不确定性", "审批版产品资料", "当场", "产品说明材料", "告知记录", "不得承诺额度、利率或审批结果", "客户要求承诺时明确拒绝", "最终结果以审批及合同为准"],
        ["SOP-LN-001-03", "二、资料受理", "客户经理", "核验身份证明并收集收入、用途和征信授权材料", "移动展业/网点", "1个工作日", "申请及证明材料", "受理清单", "只收集最小必要材料", "材料缺失一次性告知补件", "我会给您列出缺少的材料，补齐后继续办理"],
        ["SOP-LN-001-04", "二、资料受理", "客户经理", "核对复印件与原件或可信电子材料的一致性", "影像平台", "受理当日", "原件或可信电子材料", "影像及核验标记", "不得代签、不得篡改", "真伪存疑时中止受理", "材料需要进一步核验，暂不能进入审批"],
        ["SOP-LN-001-05", "三、尽调与提交", "客户经理", "核实职业、收入、负债、用途及交易背景，形成尽调意见", "信贷系统", "2个工作日", "完整申请材料", "尽调报告", "事实与判断分开记录", "信息矛盾则补充核实", "部分信息需要进一步确认，请您补充说明"],
        ["SOP-LN-001-06", "三、尽调与提交", "客户经理", "完成系统录入和双人复核后提交审批", "信贷系统", "尽调完成当日", "尽调报告及附件", "审批任务", "录入信息与材料一致", "系统校验失败则退回修正", "资料已提交，审批结果以系统通知为准"],
        ["SOP-LN-001-07", "四、审批反馈", "审批人员", "独立审查并记录通过、补件或拒绝结论", "信贷审批系统", "3个工作日", "审批任务", "审批结论", "审批独立性", "补件退回客户经理", "如需补充资料，我们会一次性联系您"],
        ["SOP-LN-001-08", "四、审批反馈", "客户经理", "使用批准口径向客户反馈结果，不披露内部评分模型", "客户联络平台", "收到结论后1个工作日", "审批结论", "通知记录", "保护内部风险规则", "客户异议按申诉流程处理", "审批综合考虑多项因素，具体以正式结果为准"],
        ["SOP-LN-001-09", "五、签约放款", "客户经理/运营", "完成合同讲解、客户本人签署、放款条件核验", "签约与放款系统", "条件满足后办理", "合同及放款条件材料", "合同、核验记录", "客户真实意愿；关键条款充分提示", "条件未满足不得放款", "请确认合同金额、期限、利率和还款安排"],
        ["SOP-LN-001-10", "六、贷后服务", "客户经理", "提示还款安排，受理信息变更、提前还款和困难客户诉求", "贷后管理平台", "持续", "存量贷款信息", "服务记录", "不得建议借新还旧规避风险", "还款困难转贷后管理人员", "如预计还款困难，请尽早联系银行协商合规方案"],
    ]
    for row in rows: ws.append(row)
    end = 4 + len(rows); style_body(ws, 5, end, 11); merge_same_values(ws, 2, 5, end)
    for col, width in enumerate([19, 16, 18, 46, 22, 18, 31, 25, 31, 30, 38], 1): ws.column_dimensions[get_column_letter(col)].width = width
    for r in range(5, end + 1): ws.row_dimensions[r].height = 72
    ws.freeze_panes = "D5"; ws.auto_filter.ref = f"A4:K{end}"; ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape"; ws.page_setup.fitToWidth = 1; ws.sheet_properties.pageSetUpPr.fitToPage = True


def add_exception_sheet(wb):
    ws = wb.create_sheet("异常处置矩阵")
    set_title(ws, "客户经理异常场景分级处置矩阵（内部参考）", 12,
              "红色为必须立即升级的高风险情形；所有时效和内部岗位名称均为演示口径。")
    for rng, value in [("A3:C3", "场景识别"), ("D3:F3", "风险判断"), ("G3:I3", "处置动作"), ("J3:L3", "沟通与留痕")]:
        ws.merge_cells(rng); ws[rng.split(":")[0]] = value
    style_header(ws, 3, 1, 12, NAVY)
    headers = ["项目", "业务领域", "异常场景", "风险等级", "识别信号", "禁止动作", "首要动作", "升级对象", "处置时限(内部参考)", "客户沟通要点", "留痕材料", "关联SOP编号"]
    ws.append(headers); style_header(ws, 4, 1, 12, BLUE)
    rows = [
        ["EX-001", "账户/反欺诈", "客户声称刚遭遇诈骗且已转账", "高", "持续通话受控、要求再次转账、提供可疑收款账户", "不得保证追回；不得让客户继续与骗子沟通", "停止转账、保护账户、保存证据并报警", "反欺诈专岗/运营主管", "立即", "说明止付结果取决于资金状态", "交易流水、客户陈述、处置时间线", "SOP-FR-002"],
        ["EX-002", "反洗钱", "客户拒绝说明突增资金来源", "高", "交易与职业收入明显不匹配、解释反复", "不得提示拆分或透露监测规则", "记录客观事实并提交预警核查", "网点反洗钱岗", "当日", "仅说明银行需履行身份识别义务", "问询记录、客户材料、系统预警", "SOP-AML-001"],
        ["EX-003", "贷款", "贷款用途材料疑似伪造", "高", "合同要素矛盾、印章异常、交易对手无法核实", "不得帮助修改或美化材料", "中止受理并保全材料", "信贷主管/合规人员", "立即", "说明材料需进一步核验", "原始材料、核验记录", "SOP-LN-001"],
        ["EX-004", "财富管理", "客户风险等级与目标产品不匹配", "中高", "系统适当性提示或拦截", "不得诱导客户重新测评提高等级", "按系统要求提示、确认或停止销售", "财富主管", "交易前", "解释风险匹配是保护客户的重要要求", "测评、提示和确认记录", "SOP-WM-001"],
        ["EX-005", "隐私保护", "客户资料误发至错误联系人", "高", "身份证、账户或交易信息已外发", "不得私下删除后隐瞒", "立即停止传播并启动信息事件报告", "信息安全/消保联系人", "立即", "未经批准不擅自对外定性或承诺", "发送记录、涉及数据清单、补救动作", "SOP-PR-001"],
        ["EX-006", "投诉服务", "客户在网点情绪激动并威胁曝光", "中高", "高声争执、直播、聚集苗头", "不得争辩、刺激或强行删除客户内容", "安排独立区域沟通并登记诉求", "网点负责人/消保岗", "立即", "先确认诉求和安全，再说明处理安排", "现场记录、投诉编号、影像（合规时）", "SOP-CS-001"],
        ["EX-007", "账户服务", "家属要求查询客户账户余额", "中", "仅提供亲属关系，无本人授权", "不得口头或截图披露客户信息", "核验有效授权或法律依据", "运营主管", "当场", "亲属关系本身不构成账户查询授权", "查询申请及授权材料", "SOP-PR-001"],
        ["EX-008", "特殊服务", "高龄客户疑似被陪同人员胁迫转账", "高", "陪同人员代答、催促、阻止单独沟通", "不得在真实意愿不明时继续办理", "分离沟通、核验意愿并启动风险处置", "网点负责人/反欺诈专岗", "立即", "以保护客户资金安全为由开展额外核实", "问询记录、风险处置记录", "SOP-OP-001"],
    ]
    for row in rows: ws.append(row)
    end = 4 + len(rows); style_body(ws, 5, end, 12)
    for r in range(5, end + 1):
        risk = ws.cell(r, 4).value
        fill = LIGHT_RED if risk == "高" else LIGHT_GOLD if risk == "中高" else LIGHT_GREEN
        ws.cell(r, 4).fill = PatternFill("solid", fgColor=fill)
        ws.cell(r, 4).font = Font(name="微软雅黑", size=10, bold=True, color=RED if risk == "高" else TEXT)
        ws.cell(r, 4).alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[r].height = 76
    for col, width in enumerate([13, 18, 30, 12, 34, 34, 36, 24, 18, 38, 34, 16], 1): ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "D5"; ws.auto_filter.ref = f"A4:L{end}"; ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape"; ws.page_setup.fitToWidth = 1; ws.sheet_properties.pageSetUpPr.fitToPage = True


def add_readme_sheet(wb):
    ws = wb.create_sheet("使用说明", 0)
    set_title(ws, "RAG Excel 小样使用说明", 8, "本工作簿用于验证复杂业务表格在现有向量化链路中的解析效果。")
    sections = [
        (4, "一、数据边界", "“华辰银行”为虚构机构。除行业通行原则外，客户等级、年龄、办理时效、系统名称及审批岗位均为内部参考，不应用于真实客户服务。"),
        (7, "二、工作簿结构", "客户经理QA：31条高频问答；准入与材料矩阵：两个子表；办理流程SOP：10步贷款流程；异常处置矩阵：8类典型异常。"),
        (10, "三、复杂格式", "包含标题跨列合并、两级/三级横向表头、同一Sheet多个子表、纵向合并业务分类、长文本自动换行、筛选和冻结窗格。"),
        (13, "四、RAG设计", "最底层表头仍采用完整语义列名；每条记录的“项目”列保持非空；问题和答案尽量自包含。这样既保留真实表格形态，也降低现有按行解析器丢失语义的概率。"),
        (16, "五、推荐测试问题", "1. 白金级客户资产门槛是多少？\n2. 消费贷款需要什么用途材料？\n3. 客户贷款用途材料疑似伪造怎么办？\n4. 客户被骗转账后第一步做什么？\n5. 家属可以查询客户余额吗？"),
    ]
    for row, title, body in sections:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        ws.cell(row, 1, title); ws.cell(row, 1).font = Font(name="微软雅黑", size=11, bold=True, color=WHITE); ws.cell(row, 1).fill = PatternFill("solid", fgColor=BLUE); ws.cell(row, 1).alignment = Alignment(vertical="center")
        ws.merge_cells(start_row=row, start_column=3, end_row=row + 1, end_column=8)
        ws.cell(row, 3, body); ws.cell(row, 3).font = Font(name="微软雅黑", size=10, color=TEXT); ws.cell(row, 3).fill = PatternFill("solid", fgColor="F8FAFC"); ws.cell(row, 3).alignment = Alignment(vertical="center", wrap_text=True); ws.cell(row, 3).border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
        section_height = 58 if row == 16 else 34
        ws.row_dimensions[row].height = section_height; ws.row_dimensions[row + 1].height = section_height
    for col, width in enumerate([18, 6, 21, 21, 21, 21, 21, 21], 1): ws.column_dimensions[get_column_letter(col)].width = width
    ws.sheet_view.showGridLines = False


def build():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    add_qa_sheet(wb)
    add_matrix_sheet(wb)
    add_sop_sheet(wb)
    add_exception_sheet(wb)
    add_readme_sheet(wb)
    wb.properties.creator = "OpenAI Codex"
    wb.properties.title = "华辰银行零售客户经理知识库小样"
    wb.properties.subject = "RAG Excel复杂表格解析测试"
    wb.save(OUTPUT_FILE)

    # Structural verification after reload.
    check = load_workbook(OUTPUT_FILE, data_only=False)
    assert check.sheetnames == ["使用说明", "客户经理QA", "准入与材料矩阵", "办理流程SOP", "异常处置矩阵"]
    assert check["客户经理QA"].max_row == 4 + len(qa_rows)
    assert len(check["客户经理QA"].merged_cells.ranges) >= 10
    assert len(check["准入与材料矩阵"].merged_cells.ranges) >= 10
    assert len(check["办理流程SOP"].merged_cells.ranges) >= 8
    assert len(check["异常处置矩阵"].merged_cells.ranges) >= 5
    print(OUTPUT_FILE)
    for ws in check.worksheets:
        print(ws.title, ws.max_row, ws.max_column, len(ws.merged_cells.ranges))


if __name__ == "__main__":
    build()
