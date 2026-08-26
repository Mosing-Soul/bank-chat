import json
import math
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


ROOT = Path(r"D:\JetBrains\project\bank-chat")
DATA = ROOT / "outputs" / "rag_excel_sample_work" / "sop_documents.json"
TMP = ROOT / "outputs" / "rag_excel_sample_work" / "sop_sources"

NAVY = "17365D"
BLUE = "2F75B5"
LIGHT_BLUE = "D9EAF7"
PALE = "F7F9FC"
GOLD = "FFF2CC"
RED = "9C0006"
WHITE = "FFFFFF"
TEXT = "1F2937"
THIN = Side(style="thin", color="C9D2DC")


def merge_row(ws, row, text, start=1, end=10, height=26, fill=None, font=None, align=None):
    ws.merge_cells(start_row=row, start_column=start, end_row=row, end_column=end)
    cell = ws.cell(row, start, text)
    cell.alignment = align or Alignment(vertical="center", wrap_text=True)
    cell.font = font or Font(name="微软雅黑", size=10, color=TEXT)
    if fill:
        cell.fill = PatternFill("solid", fgColor=fill)
    ws.row_dimensions[row].height = height
    return cell


def add_table(ws, row, table):
    headers = table["headers"]
    data = table["rows"]
    n = len(headers)
    spans = []
    for idx in range(n):
        start = math.floor(idx * 10 / n) + 1
        end = math.floor((idx + 1) * 10 / n)
        spans.append((start, max(start, end)))
    for idx, header in enumerate(headers):
        start, end = spans[idx]
        ws.merge_cells(start_row=row, start_column=start, end_row=row, end_column=end)
        c = ws.cell(row, start, header)
        c.font = Font(name="微软雅黑", size=9, bold=True, color=WHITE)
        c.fill = PatternFill("solid", fgColor=BLUE)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col in range(1, 11):
        ws.cell(row, col).border = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)
    ws.row_dimensions[row].height = 28
    row += 1
    for ridx, values in enumerate(data):
        max_chars = 0
        for idx, value in enumerate(values):
            start, end = spans[idx]
            ws.merge_cells(start_row=row, start_column=start, end_row=row, end_column=end)
            c = ws.cell(row, start, value)
            c.font = Font(name="微软雅黑", size=8.5, color=TEXT)
            c.fill = PatternFill("solid", fgColor=PALE if ridx % 2 else WHITE)
            c.alignment = Alignment(vertical="top", wrap_text=True)
            max_chars = max(max_chars, len(str(value)) / max(1, end - start + 1))
        for col in range(1, 11):
            ws.cell(row, col).border = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)
        ws.row_dimensions[row].height = min(96, max(34, 18 + max_chars * 2.2))
        row += 1
    return row + 1


def build_doc(item, index):
    wb = Workbook()
    ws = wb.active
    ws.title = "SOP正文"
    ws.sheet_view.showGridLines = False
    for col in range(1, 11):
        ws.column_dimensions[chr(64 + col)].width = 13.5
    row = 1
    merge_row(ws, row, item["title"], height=38, fill=NAVY,
              font=Font(name="微软雅黑", size=19, bold=True, color=WHITE),
              align=Alignment(horizontal="center", vertical="center")); row += 1
    merge_row(ws, row, item["subtitle"], height=26, fill=LIGHT_BLUE,
              font=Font(name="微软雅黑", size=11, italic=True, color=NAVY),
              align=Alignment(horizontal="center", vertical="center")); row += 2
    meta = {
        "headers": item["metaHeaders"],
        "rows": [[item["code"], item["version"], item["effectiveDate"], item["owner"]]],
    }
    row = add_table(ws, row, meta)
    merge_row(ws, row, item["scopeLabel"] + item["scope"], height=42, fill=PALE,
              font=Font(name="微软雅黑", size=10, bold=True, color=TEXT)); row += 2
    merge_row(ws, row, item["warning"], height=54, fill=GOLD,
              font=Font(name="微软雅黑", size=10, bold=True, color=RED)); row += 2

    for section in item["sections"]:
        if section.get("pageBreakBefore") and row > 1:
            ws.row_breaks.append(__import__("openpyxl").worksheet.pagebreak.Break(id=row - 1))
        merge_row(ws, row, section["heading"], height=28, fill=NAVY,
                  font=Font(name="微软雅黑", size=13, bold=True, color=WHITE)); row += 1
        for paragraph in section.get("paragraphs", []):
            height = min(80, max(34, 20 + len(paragraph) * 0.55))
            merge_row(ws, row, paragraph, height=height); row += 1
        for bullet in section.get("bullets", []):
            height = min(70, max(28, 18 + len(bullet) * 0.45))
            merge_row(ws, row, "• " + bullet, start=1, end=10, height=height,
                      fill=PALE if row % 2 else WHITE); row += 1
        if section.get("table"):
            row = add_table(ws, row, section["table"])
        else:
            row += 1

    ws.freeze_panes = "A4"
    ws.print_area = f"A1:J{row - 1}"
    ws.print_options.horizontalCentered = True
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.oddHeader.center.text = f"&B{item['code']} | {item['version']}"
    ws.oddHeader.center.size = 8
    ws.oddFooter.center.text = "Mock knowledge document | Page &P of &N"
    ws.oddFooter.center.size = 8
    ws.sheet_properties.pageSetUpPr.autoPageBreaks = True
    path = TMP / f"sop_{index}.xlsx"
    wb.save(path)
    return path


def main():
    TMP.mkdir(parents=True, exist_ok=True)
    data = json.loads(DATA.read_text(encoding="utf-8"))
    mapping = []
    for index, item in enumerate(data["documents"], 1):
        source = build_doc(item, index)
        mapping.append({"source": str(source), "pdf": item["filename"]})
    (TMP / "mapping.json").write_text(json.dumps(mapping, ensure_ascii=False, indent=2), encoding="utf-8-sig")
    print("\n".join(str(x["source"]) for x in mapping))


if __name__ == "__main__":
    main()
