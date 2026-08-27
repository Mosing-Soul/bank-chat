from pathlib import Path
import json
import fitz
from openpyxl import load_workbook

ROOT=Path(r"D:\JetBrains\project\bank-chat")
ASSETS=ROOT/"bank-agent-demo"/"assets"
PRE=ROOT/"outputs"/"rag_expansion_v2"/"pdf_previews"
PRE.mkdir(parents=True,exist_ok=True)

xlsx=ASSETS/"华辰银行零售客户经理业务知识问答手册_V2.0.xlsx"
wb=load_workbook(xlsx,data_only=False)
ws=wb["客户经理QA"]
qa_count=ws.max_row-4
questions=[ws.cell(r,4).value for r in range(5,ws.max_row+1)]
assert qa_count==150 and len(set(questions))==150
assert len(ws.merged_cells.ranges)>=10

results=[]
for pdf in sorted(ASSETS.glob("华辰银行*.pdf")):
    doc=fitz.open(pdf)
    chars=sum(len(p.get_text()) for p in doc)
    assert len(doc)>=4 and chars>1000
    for i,p in enumerate(doc):
        pix=p.get_pixmap(matrix=fitz.Matrix(0.75,0.75),alpha=False)
        pix.save(PRE/f"{pdf.stem}_p{i+1:02d}.png")
    results.append({"file":pdf.name,"pages":len(doc),"text_chars":chars})
assert len(results)==10
report={"qa_count":qa_count,"unique_questions":len(set(questions)),"qa_merges":len(ws.merged_cells.ranges),"sop_count":len(results),"sops":results}
(ROOT/"outputs"/"rag_expansion_v2"/"verification.json").write_text(json.dumps(report,ensure_ascii=False,indent=2),encoding="utf-8")
print(json.dumps(report,ensure_ascii=False,indent=2))
